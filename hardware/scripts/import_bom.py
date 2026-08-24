#!/usr/bin/env python3
"""
Meridian Ecology — BOM Import Script
Reads Sensor_Species_Family___Full_Bill_of_Materials.xlsx and populates:
  - bom_sources
  - parts
  - bom_line_items
  - inventory
  - node_configurations + node_parts  (garden field nodes from session)
  - deployed_nodes                    (scan coordinates from Scaniverse 2026-05-13)

Target: PostgreSQL 17 on Inferno (192.168.0.28)
Usage:
  python3 import_bom.py [--xlsx /path/to/BOM.xlsx] [--dry-run]
  env vars: PGHOST PGPORT PGDATABASE PGUSER PGPASSWORD
"""

import os
import sys
import argparse
import uuid
from pathlib import Path
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values, RealDictCursor

# ── Config ──────────────────────────────────────────────────────────────────

DEFAULT_XLSX = Path(__file__).parent / "Sensor_Species_Family___Full_Bill_of_Materials.xlsx"

DB = dict(
    host     = os.environ.get("PGHOST",     "192.168.0.28"),
    port     = int(os.environ.get("PGPORT", "5432")),
    dbname   = os.environ.get("PGDATABASE", "sensor_ecology"),
    user     = os.environ.get("PGUSER",     "postgres"),
    password = os.environ.get("PGPASSWORD", ""),
)

SSF_BOM_NAME     = "Sensor Species Family"
GARDEN_BOM_NAME  = "Garden Field Nodes"
MERIDIAN_BOM_NAME = "Meridian Core Infrastructure"

# ── Sensor interface / I2C address lookup ────────────────────────────────────
# Keyed on substrings that appear in component name or description.
# First match wins; order matters for specificity.

INTERFACE_HINTS = [
    # (substring_in_name_lower, interface_type, i2c_address, voltage_v)
    ("bme688",        "I2C",  "0x76",  3.3),
    ("bme280",        "I2C",  "0x76",  3.3),
    ("tsl2591",       "I2C",  "0x29",  3.3),
    ("apds-9960",     "I2C",  "0x39",  3.3),
    ("apds9960",      "I2C",  "0x39",  3.3),
    ("scd41",         "I2C",  "0x62",  3.3),
    ("veml6075",      "I2C",  "0x10",  3.3),
    ("sen55",         "I2C",  "0x69",  3.3),
    ("adxl345",       "I2C",  "0x53",  3.3),
    ("ld2410",        "UART", None,    3.3),
    ("pms5003",       "UART", None,    5.0),
    ("inmp441",       "I2S",  None,    3.3),
    ("max4466",       "analog", None,  3.3),
    ("max9814",       "analog", None,  3.3),
    ("pir",           "digital", None, 5.0),
    ("hc-sr501",      "digital", None, 5.0),
    ("ws2812b",       "digital", None, 5.0),
    ("neopixel",      "digital", None, 5.0),
    ("piezo",         "analog", None,  3.3),
    ("pico w",        "USB",  None,    5.0),
    ("qt py",         "USB",  None,    5.0),
    ("rpi 4b",        "USB",  None,    5.0),
    ("raspberry pi 4","USB",  None,    5.0),
    ("raspberry pi pico","USB", None,  5.0),
    ("tp4056",        "power", None,   5.0),
    ("jst",           "connector", None, None),
    ("molex",         "connector", None, None),
    ("rj45",          "connector", None, None),
    ("usb-c",         "power", None,   5.0),
    ("buck converter","power", None,   None),
    ("lipo",          "power", None,   3.7),
    ("psu",           "power", None,   5.0),
    ("power supply",  "power", None,   5.0),
]

def detect_interface(name: str, description: str):
    text = (name + " " + (description or "")).lower()
    for hint, iface, addr, volt in INTERFACE_HINTS:
        if hint in text:
            return iface, addr, volt
    return None, None, None

# ── Print material detection ─────────────────────────────────────────────────

def detect_print_material(category: str, material_spec: str) -> str | None:
    if category != "3D Printed Parts":
        return None
    spec = (material_spec or "").upper()
    for mat in ["PETG-CF", "TPU 95A", "TPU 85A", "PETG", "ASA", "PLA", "ABS"]:
        if mat in spec:
            return mat
    return None

def detect_print_mass(material_spec: str) -> float | None:
    import re
    if not material_spec:
        return None
    m = re.search(r"~(\d+(?:\.\d+)?)g", material_spec)
    return float(m.group(1)) if m else None

# ── Priority / status normalisation ─────────────────────────────────────────

PRIORITY_MAP = {
    "critical":    "critical",
    "recommended": "recommended",
    "optional":    "optional",
}

STATUS_MAP = {
    "to order":  "to_order",
    "in stock":  "in_stock",
    "ordered":   "ordered",
    "received":  "received",
}

def norm_priority(raw: str) -> str:
    return PRIORITY_MAP.get((raw or "").strip().lower(), "recommended")

def norm_status(raw: str) -> str:
    return STATUS_MAP.get((raw or "").strip().lower(), "to_order")

# ── Supplier parsing ─────────────────────────────────────────────────────────

def parse_suppliers(raw: str) -> list[str]:
    if not raw:
        return []
    return [s.strip() for s in raw.replace(" / ", "/").split("/") if s.strip()]

# ── Phase 1: read xlsx ───────────────────────────────────────────────────────

def read_bom(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb["BOM"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        item_no = row[0]
        if not isinstance(item_no, (int, float)):
            continue
        name        = str(row[1]).strip() if row[1] else ""
        description = str(row[2]).strip() if row[2] else ""
        category    = str(row[3]).strip() if row[3] else ""
        morphology  = str(row[4]).strip() if row[4] else ""
        material    = str(row[5]).strip() if row[5] else ""
        qty         = float(row[6]) if row[6] else 0
        unit        = str(row[7]).strip() if row[7] else "ea"
        unit_cost   = float(row[8]) if isinstance(row[8], (int, float)) else None
        # row[9] is formula string in xlsx — compute ourselves
        supplier    = str(row[10]).strip() if row[10] else ""
        priority    = str(row[11]).strip() if row[11] else "Recommended"
        status      = str(row[12]).strip() if row[12] else "To Order"

        if not name:
            continue

        rows.append({
            "item_no":    int(item_no),
            "name":       name,
            "description": description,
            "category":   category,
            "morphology": morphology,
            "material":   material,
            "qty":        qty,
            "unit":       unit,
            "unit_cost":  unit_cost,
            "total_cost": round(qty * unit_cost, 4) if unit_cost else None,
            "suppliers":  parse_suppliers(supplier),
            "priority":   norm_priority(priority),
            "status":     norm_status(status),
        })
    return rows

# ── Phase 2: seed bom_sources ────────────────────────────────────────────────

BOM_SOURCES = [
    (SSF_BOM_NAME,
     "Kinetic sculptural installation — Whitehorse YT. 20 primary components across four morphological types.",
     "installation"),
    (GARDEN_BOM_NAME,
     "Outdoor sensor ecology deployment — garden plot, Whitehorse YT. Initial autonomy testing phase.",
     "field"),
    (MERIDIAN_BOM_NAME,
     "Indoor research ecology — Inferno Pi5, Sensor Pi, edge nodes, AR layer.",
     "research"),
]

def upsert_bom_sources(cur) -> dict[str, str]:
    """Returns {bom_name: uuid}"""
    result = {}
    for name, desc, ptype in BOM_SOURCES:
        cur.execute("""
            INSERT INTO bom_sources (id, name, description, project_type)
            VALUES (uuid_generate_v4(), %s, %s, %s)
            ON CONFLICT (name) DO UPDATE SET description = EXCLUDED.description
            RETURNING id, name
        """, (name, desc, ptype))
        row = cur.fetchone()
        result[row["name"]] = str(row["id"])
    return result

# ── Phase 3: upsert parts ────────────────────────────────────────────────────

def upsert_parts(cur, rows: list[dict]) -> dict[str, str]:
    """Returns {part_name: uuid}"""
    result = {}
    for r in rows:
        iface, addr, volt = detect_interface(r["name"], r["description"])
        print_mat  = detect_print_material(r["category"], r["material"])
        print_mass = detect_print_mass(r["material"])

        cur.execute("""
            INSERT INTO parts (
                id, name, description, category, subcategory,
                material_spec, unit_type, unit_cost_cad,
                supplier_options, interface_type, i2c_address,
                voltage_v, print_material, print_mass_g
            ) VALUES (
                uuid_generate_v4(), %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s
            )
            ON CONFLICT (name) DO UPDATE SET
                description     = EXCLUDED.description,
                material_spec   = EXCLUDED.material_spec,
                unit_cost_cad   = EXCLUDED.unit_cost_cad,
                supplier_options = EXCLUDED.supplier_options,
                interface_type  = EXCLUDED.interface_type,
                i2c_address     = EXCLUDED.i2c_address,
                voltage_v       = EXCLUDED.voltage_v,
                print_material  = EXCLUDED.print_material,
                print_mass_g    = EXCLUDED.print_mass_g,
                updated_at      = NOW()
            RETURNING id, name
        """, (
            r["name"], r["description"], r["category"], r.get("subcategory"),
            r["material"], r["unit"], r["unit_cost"],
            r["suppliers"] or None, iface, addr,
            volt, print_mat, print_mass,
        ))
        row = cur.fetchone()
        result[row["name"]] = str(row["id"])
    return result

# ── Phase 4: upsert inventory ────────────────────────────────────────────────

def upsert_inventory(cur, rows: list[dict], part_ids: dict[str, str]):
    for r in rows:
        pid = part_ids.get(r["name"])
        if not pid:
            continue
        qty_on_hand = r["qty"] if r["status"] == "in_stock" else 0
        qty_on_order = r["qty"] if r["status"] == "ordered" else 0
        cur.execute("""
            INSERT INTO inventory (id, part_id, qty_on_hand, qty_on_order, status)
            VALUES (uuid_generate_v4(), %s, %s, %s, %s)
            ON CONFLICT (part_id) DO UPDATE SET
                qty_on_hand  = EXCLUDED.qty_on_hand,
                qty_on_order = EXCLUDED.qty_on_order,
                status       = EXCLUDED.status,
                last_updated = NOW()
        """, (pid, qty_on_hand, qty_on_order, r["status"]))

# ── Phase 5: upsert bom_line_items ───────────────────────────────────────────

def upsert_bom_line_items(cur, rows: list[dict], bom_id: str, part_ids: dict[str, str]):
    for r in rows:
        pid = part_ids.get(r["name"])
        if not pid:
            continue
        cur.execute("""
            INSERT INTO bom_line_items (
                id, bom_id, part_id, source_item_no,
                qty_required, morphology, priority, notes
            ) VALUES (
                uuid_generate_v4(), %s, %s, %s,
                %s, %s, %s, %s
            )
            ON CONFLICT (bom_id, part_id, morphology) DO UPDATE SET
                qty_required = EXCLUDED.qty_required,
                priority     = EXCLUDED.priority
        """, (
            bom_id, pid, r["item_no"],
            r["qty"], r["morphology"], r["priority"],
            r.get("notes"),
        ))

# ── Phase 6: seed node configurations ───────────────────────────────────────
# Garden field nodes defined during session 2026-05-15

NODE_CONFIGS = [
    {
        "name": "Hub Gateway",
        "slug": "hub-gateway",
        "description": "Central coordinator for garden field ecology. No sensing — MQTT relay, OTA management, local SQLite state cache.",
        "node_family": "garden_field",
        "enclosure_type": "shed_mounted",
        "deployment_context": "shed_exterior",
        "controller": "pi_pico_w",
        "comms_protocol": ["mqtt", "wifi"],
        "version": "0.1",
    },
    {
        "name": "E1 Central Environmental",
        "slug": "e1-central",
        "description": "Core ecology sensing node in open garden area. Richest sensor package — microclimate, air quality, UV, presence, acoustic.",
        "node_family": "garden_field",
        "enclosure_type": "puck_outdoor",
        "deployment_context": "outdoor_exposed",
        "controller": "pi_pico_w",
        "comms_protocol": ["mqtt", "wifi"],
        "version": "0.1",
    },
    {
        "name": "B1 Bed North",
        "slug": "b1-bed-north",
        "description": "Garden bed sensor node, north position. Soil moisture + local temp + light for plant-relevant monitoring.",
        "node_family": "garden_field",
        "enclosure_type": "puck_outdoor",
        "deployment_context": "outdoor_garden_bed",
        "controller": "pi_pico_w",
        "comms_protocol": ["mqtt", "wifi"],
        "version": "0.1",
    },
    {
        "name": "B2 Bed South",
        "slug": "b2-bed-south",
        "description": "Garden bed sensor node, south position. Soil moisture + local temp + light.",
        "node_family": "garden_field",
        "enclosure_type": "puck_outdoor",
        "deployment_context": "outdoor_garden_bed",
        "controller": "pi_pico_w",
        "comms_protocol": ["mqtt", "wifi"],
        "version": "0.1",
    },
    {
        "name": "W Sky Reference",
        "slug": "w-sky-reference",
        "description": "Elevated, unobstructed sky reference node. NE corner, open sky view. All-atmosphere baseline — everything else calibrates against this.",
        "node_family": "garden_field",
        "enclosure_type": "puck_outdoor",
        "deployment_context": "outdoor_elevated",
        "controller": "pi_pico_w",
        "comms_protocol": ["mqtt", "wifi"],
        "version": "0.1",
    },
    {
        "name": "P Entry Presence",
        "slug": "p-entry-presence",
        "description": "Presence and approach detection at garden entry. Radar + PIR combination for false-positive reduction.",
        "node_family": "garden_field",
        "enclosure_type": "puck_outdoor",
        "deployment_context": "outdoor_entry",
        "controller": "pi_pico_w",
        "comms_protocol": ["mqtt", "wifi"],
        "version": "0.1",
    },
]

def upsert_node_configs(cur) -> dict[str, str]:
    """Returns {slug: uuid}"""
    result = {}
    for nc in NODE_CONFIGS:
        cur.execute("""
            INSERT INTO node_configurations (
                id, name, slug, description, node_family,
                enclosure_type, deployment_context, controller,
                comms_protocol, version
            ) VALUES (
                uuid_generate_v4(), %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s
            )
            ON CONFLICT (slug) DO UPDATE SET
                description        = EXCLUDED.description,
                deployment_context = EXCLUDED.deployment_context
            RETURNING id, slug
        """, (
            nc["name"], nc["slug"], nc["description"], nc["node_family"],
            nc["enclosure_type"], nc["deployment_context"], nc["controller"],
            nc["comms_protocol"], nc["version"],
        ))
        row = cur.fetchone()
        result[row["slug"]] = str(row["id"])
    return result

# ── Phase 7: seed node_parts ─────────────────────────────────────────────────
# Links node configs to parts by part name lookup.
# Parts must already exist in the parts table.

NODE_PARTS = {
    "e1-central": [
        # role, part_name_substring, qty, bus_position, required
        ("controller",      "Raspberry Pi Pico W",          1, "main",        True),
        ("primary_sensor",  "BME688",                       1, "i2c_0x76",    True),
        ("primary_sensor",  "TSL2591",                      1, "i2c_0x29",    True),
        ("primary_sensor",  "Capacitive Soil",              1, "gpio_adc",    False),
        ("enclosure",       "Pollen Pod Shell",             1, None,          False),
        ("connector",       "JST-SH 1.0mm 4-pin",          4, "i2c_bus",     True),
    ],
    "b1-bed-north": [
        ("controller",      "Raspberry Pi Pico W",          1, "main",        True),
        ("primary_sensor",  "BME280",                       1, "i2c_0x76",    True),
        ("primary_sensor",  "TSL2591",                      1, "i2c_0x29",    True),
        ("primary_sensor",  "Capacitive Soil",              1, "gpio_adc",    True),
        ("enclosure",       "Pollen Pod Shell",             1, None,          False),
        ("connector",       "JST-SH 1.0mm 4-pin",          2, "i2c_bus",     True),
    ],
    "b2-bed-south": [
        ("controller",      "Raspberry Pi Pico W",          1, "main",        True),
        ("primary_sensor",  "BME280",                       1, "i2c_0x76",    True),
        ("primary_sensor",  "TSL2591",                      1, "i2c_0x29",    True),
        ("primary_sensor",  "Capacitive Soil",              1, "gpio_adc",    True),
        ("enclosure",       "Pollen Pod Shell",             1, None,          False),
        ("connector",       "JST-SH 1.0mm 4-pin",          2, "i2c_bus",     True),
    ],
    "w-sky-reference": [
        ("controller",      "Raspberry Pi Pico W",          1, "main",        True),
        ("primary_sensor",  "BME280",                       1, "i2c_0x76",    True),
        ("primary_sensor",  "TSL2591",                      1, "i2c_0x29",    True),
        ("enclosure",       "Perch Pod Main Body",          1, None,          False),
        ("connector",       "JST-SH 1.0mm 4-pin",          3, "i2c_bus",     True),
    ],
    "p-entry-presence": [
        ("controller",      "Raspberry Pi Pico W",          1, "main",        True),
        ("primary_sensor",  "PIR Motion Sensor HC-SR501",   1, "gpio_5",      True),
        ("primary_sensor",  "APDS-9960",                    1, "i2c_0x39",    True),
        ("enclosure",       "Perch Pod Main Body",          1, None,          False),
        ("connector",       "JST-SH 1.0mm 4-pin",          2, "i2c_bus",     True),
    ],
    "hub-gateway": [
        ("controller",      "Raspberry Pi Pico W",          1, "main",        True),
        ("power",           "LiPo Battery 3.7V 2000mAh",   1, "backup",      True),
        ("power",           "LiPo Charger",                 1, "charge",      True),
        ("connector",       "USB-C Panel Mount",            1, "power_in",    True),
    ],
}

def upsert_node_parts(cur, node_config_ids: dict[str, str], part_ids: dict[str, str]):
    inserted = 0
    skipped  = 0
    for slug, parts in NODE_PARTS.items():
        nc_id = node_config_ids.get(slug)
        if not nc_id:
            print(f"  WARNING: node config not found for slug '{slug}'")
            continue
        for role, part_substr, qty, bus_pos, required in parts:
            # Find part by substring match in part_ids keys
            matches = [name for name in part_ids if part_substr.lower() in name.lower()]
            if not matches:
                print(f"  WARNING: no part matching '{part_substr}' for {slug}")
                skipped += 1
                continue
            part_name = matches[0]
            pid = part_ids[part_name]
            cur.execute("""
                INSERT INTO node_parts (
                    id, node_config_id, part_id, qty, role, bus_position, is_required
                ) VALUES (
                    uuid_generate_v4(), %s, %s, %s, %s, %s, %s
                )
                ON CONFLICT DO NOTHING
            """, (nc_id, pid, qty, role, bus_pos, required))
            inserted += 1
    return inserted, skipped

# ── Phase 8: seed deployed_nodes ─────────────────────────────────────────────
# Positions from Scaniverse scan 2026-05-13, coordinate space in metres.
# Compass bearing 212.8° (SSW), GPS 60.7038°N 135.0247°W, elev 649.5m

DEPLOYED_NODES = [
    # label, slug, location_name, scan_x, scan_y, scan_z, mqtt_prefix
    ("H",  "hub-gateway",      "shed east wall",         -3.5,  0.0,  0.0, "meridian/garden/hub"),
    ("E1", "e1-central",       "garden centre open",      1.5,  0.0,  0.0, "meridian/garden/e1"),
    ("B1", "b1-bed-north",     "north bed (TBC)",         0.0,  0.0,  1.8, "meridian/garden/b1"),
    ("B2", "b2-bed-south",     "south bed (TBC)",         0.0,  0.0, -1.8, "meridian/garden/b2"),
    ("W",  "w-sky-reference",  "NE corner elevated",      4.0,  0.5,  2.2, "meridian/garden/w"),
    ("P",  "p-entry-presence", "east entry approach",     5.5,  0.0,  0.0, "meridian/garden/p"),
]

def upsert_deployed_nodes(cur, node_config_ids: dict[str, str]):
    for label, slug, location, sx, sy, sz, mqtt in DEPLOYED_NODES:
        nc_id = node_config_ids.get(slug)
        cur.execute("""
            INSERT INTO deployed_nodes (
                id, node_config_id, label, location_name,
                scan_x, scan_y, scan_z, scan_source,
                mqtt_topic_prefix, is_active
            ) VALUES (
                uuid_generate_v4(), %s, %s, %s,
                %s, %s, %s, %s,
                %s, FALSE
            )
            ON CONFLICT (label) DO UPDATE SET
                location_name      = EXCLUDED.location_name,
                scan_x             = EXCLUDED.scan_x,
                scan_y             = EXCLUDED.scan_y,
                scan_z             = EXCLUDED.scan_z,
                mqtt_topic_prefix  = EXCLUDED.mqtt_topic_prefix
        """, (nc_id, label, location, sx, sy, sz,
              "scaniverse_2026-05-13", mqtt))

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import SSF BOM into Meridian parts DB")
    parser.add_argument("--xlsx",    default=str(DEFAULT_XLSX), help="Path to BOM xlsx")
    parser.add_argument("--dry-run", action="store_true",       help="Parse only, no DB writes")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found at {xlsx_path}")
        sys.exit(1)

    print(f"Reading BOM: {xlsx_path}")
    rows = read_bom(xlsx_path)
    print(f"  {len(rows)} line items parsed")

    # Stats
    by_status = {}
    by_priority = {}
    for r in rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
        by_priority[r["priority"]] = by_priority.get(r["priority"], 0) + 1
    print(f"  Status:   {dict(sorted(by_status.items()))}")
    print(f"  Priority: {dict(sorted(by_priority.items()))}")

    total_cost = sum(r["total_cost"] for r in rows if r["total_cost"])
    print(f"  Estimated total: CAD ${total_cost:.2f}")

    if args.dry_run:
        print("\nDry run — no database writes.")
        for r in rows[:5]:
            iface, addr, volt = detect_interface(r["name"], r["description"])
            print(f"  [{r['item_no']:3d}] {r['name'][:45]:<45} "
                  f"iface={iface or '—':8} addr={addr or '—':8} status={r['status']}")
        print("  ...")
        return

    print(f"\nConnecting to {DB['host']}:{DB['port']}/{DB['dbname']}...")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        print(f"ERROR: Could not connect to database: {e}")
        print("Hint: set PGHOST PGPORT PGDATABASE PGUSER PGPASSWORD env vars")
        sys.exit(1)

    conn.autocommit = False
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        # ── Phase 2: BOM sources
        print("\nPhase 2: Upserting BOM sources...")
        bom_ids = upsert_bom_sources(cur)
        for name, bid in bom_ids.items():
            print(f"  {name}: {bid}")

        ssf_bom_id = bom_ids[SSF_BOM_NAME]

        # ── Phase 3: Parts
        print(f"\nPhase 3: Upserting {len(rows)} parts...")
        part_ids = upsert_parts(cur, rows)
        print(f"  {len(part_ids)} parts upserted")

        # Interface detection summary
        i2c_parts = [n for n in part_ids if any(
            h[0] in n.lower() for h in INTERFACE_HINTS if h[1] == "I2C"
        )]
        print(f"  I2C sensors detected: {len(i2c_parts)}")

        # ── Phase 4: Inventory
        print("\nPhase 4: Upserting inventory states...")
        upsert_inventory(cur, rows, part_ids)
        in_stock = [r for r in rows if r["status"] == "in_stock"]
        print(f"  {len(in_stock)} parts marked in_stock")
        for r in in_stock:
            print(f"    ✓ {r['name']}")

        # ── Phase 5: BOM line items
        print(f"\nPhase 5: Upserting BOM line items for '{SSF_BOM_NAME}'...")
        upsert_bom_line_items(cur, rows, ssf_bom_id, part_ids)
        print(f"  {len(rows)} line items linked to SSF BOM")

        # ── Phase 6: Node configurations
        print("\nPhase 6: Seeding garden node configurations...")
        node_config_ids = upsert_node_configs(cur)
        for slug, nid in node_config_ids.items():
            print(f"  {slug}: {nid}")

        # ── Phase 7: Node parts
        print("\nPhase 7: Linking parts to node configurations...")
        inserted, skipped = upsert_node_parts(cur, node_config_ids, part_ids)
        print(f"  {inserted} node_parts inserted, {skipped} skipped (part not found)")

        # ── Phase 8: Deployed nodes
        print("\nPhase 8: Seeding deployed node positions (from Scaniverse 2026-05-13)...")
        upsert_deployed_nodes(cur, node_config_ids)
        print(f"  {len(DEPLOYED_NODES)} nodes placed in scan coordinate space")

        conn.commit()
        print("\n✓ Import complete. All changes committed.")

        # ── Summary query
        print("\n── Summary ──────────────────────────────────────────────────")
        cur.execute("SELECT COUNT(*) AS n FROM parts")
        print(f"  parts:            {cur.fetchone()['n']}")
        cur.execute("SELECT COUNT(*) AS n FROM bom_line_items")
        print(f"  bom_line_items:   {cur.fetchone()['n']}")
        cur.execute("SELECT status, COUNT(*) AS n FROM inventory GROUP BY status ORDER BY n DESC")
        for row in cur.fetchall():
            print(f"  inventory [{row['status']:12}]: {row['n']}")
        cur.execute("SELECT COUNT(*) AS n FROM node_configurations")
        print(f"  node_configs:     {cur.fetchone()['n']}")
        cur.execute("SELECT COUNT(*) AS n FROM node_parts")
        print(f"  node_parts:       {cur.fetchone()['n']}")
        cur.execute("SELECT COUNT(*) AS n FROM deployed_nodes")
        print(f"  deployed_nodes:   {cur.fetchone()['n']}")
        cur.execute("""
            SELECT label, location_name, scan_x, scan_z,
                   COALESCE(mqtt_topic_prefix, '—') AS mqtt
            FROM deployed_nodes ORDER BY label
        """)
        print("\n  Deployed node positions:")
        for row in cur.fetchall():
            print(f"    {row['label']:4} ({row['scan_x']:+.1f}, {row['scan_z']:+.1f})m  "
                  f"{row['location_name']:<30}  {row['mqtt']}")
        print("─────────────────────────────────────────────────────────────")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
