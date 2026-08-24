#!/usr/bin/env python3
"""
Meridian v2 — general BOM importer.

Replaces hardware/scripts/import_bom.py, which only ever worked for one
xlsx (Sensor Species Family's "BOM" sheet) and pointed at v1's sensor_ecology
db on native Postgres (192.168.0.28:5432). This targets the v2 `meridian`
db (pgvector container, port 5544) and handles all four current BOM sources,
each with its own real column layout (verified against the actual files,
not assumed):

  wrist_puck      hardware/bom/wrist_puck_bom_v1_DRAFT.csv   (flat CSV, CAD)
  ssf             Sensor Species Family — Full Bill of Materials.xlsx, sheet "BOM"
  morphogenesis   Morphogenesis BOM ....xlsx, sheets "Hardware BOM" + "CF Upgrade Kit"
  lora_backbone   Portenta–RPi5–Jetson LoRa Backbone BOM.xlsx, 5 category sheets, USD

lora_backbone prices are converted USD -> CAD at import time (see FX_RATE
below) since schemas/003's parts.unit_cost_cad assumes one currency. The
rate and its source/date are recorded in each such part's notes for
auditability — it is an estimate, not a locked price.

Idempotent: every upsert is keyed (parts.name — see schemas/004 for the
UNIQUE constraint this depends on; bom_sources.name; bom_line_items'
(bom_id, part_id, morphology) unique triple), so re-running after fixing
a row in the source file updates in place rather than duplicating.

Usage:
  python3 import_bom_v2.py --source all --dry-run
  python3 import_bom_v2.py --source wrist_puck
  DATABASE_URL=postgresql://meridian:...@localhost:5544/meridian python3 import_bom_v2.py --source all

Env:
  DATABASE_URL, or PGHOST/PGPORT/PGDATABASE/PGUSER/PGPASSWORD
  (defaults match deploy/meridian.env: v2 meridian db, NOT v1 sensor_ecology)
"""

import os
import sys
import csv
import argparse
from pathlib import Path

from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import RealDictCursor

# ── FX ────────────────────────────────────────────────────────────────────
# 1 USD = 1.38595 CAD, xe.com, 2026-08-24 17:26 UTC. Re-check before a real
# purchase — this is an estimate for BOM planning, not a locked rate.
FX_USD_TO_CAD = 1.38595
FX_SOURCE_NOTE = "USD->CAD @ 1.38595 (xe.com, 2026-08-24)"

# ── DB ────────────────────────────────────────────────────────────────────

def db_connect():
    url = os.environ.get("DATABASE_URL")
    if url:
        return psycopg2.connect(url, cursor_factory=RealDictCursor)
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "192.168.0.28"),
        port=int(os.environ.get("PGPORT", "5544")),
        dbname=os.environ.get("PGDATABASE", "meridian"),
        user=os.environ.get("PGUSER", "meridian"),
        password=os.environ.get("PGPASSWORD", ""),
        cursor_factory=RealDictCursor,
    )

# ── shared interface/I2C detection (from the old script, proven) ───────────

INTERFACE_HINTS = [
    ("bme688", "I2C", "0x76", 3.3), ("bme280", "I2C", "0x76", 3.3),
    ("tsl2591", "I2C", "0x29", 3.3), ("apds-9960", "I2C", "0x39", 3.3),
    ("apds9960", "I2C", "0x39", 3.3), ("drv2605", "I2C", "0x5A", 3.3),
    ("scd41", "I2C", "0x62", 3.3), ("veml6075", "I2C", "0x10", 3.3),
    ("sen55", "I2C", "0x69", 3.3), ("adxl345", "I2C", "0x53", 3.3),
    ("ld2410", "UART", None, 3.3), ("pms5003", "UART", None, 5.0),
    ("inmp441", "I2S", None, 3.3), ("max4466", "analog", None, 3.3),
    ("max9814", "analog", None, 3.3), ("pir", "digital", None, 5.0),
    ("hc-sr501", "digital", None, 5.0), ("ws2812b", "digital", None, 5.0),
    ("neopixel", "digital", None, 5.0), ("piezo", "analog", None, 3.3),
    ("xiao esp32", "USB", None, 3.3), ("pico w", "USB", None, 5.0),
    ("qt py", "USB", None, 5.0), ("rpi 4b", "USB", None, 5.0),
    ("raspberry pi 5", "USB", None, 5.0), ("raspberry pi 4", "USB", None, 5.0),
    ("raspberry pi pico", "USB", None, 5.0), ("portenta", "USB", None, 5.0),
    ("jetson", "USB", None, 5.0), ("tp4056", "power", None, 5.0),
    ("jst", "connector", None, None), ("molex", "connector", None, None),
    ("rj45", "connector", None, None), ("usb-c", "power", None, 5.0),
    ("buck converter", "power", None, None), ("lipo", "power", None, 3.7),
    ("psu", "power", None, 5.0), ("power supply", "power", None, 5.0),
    ("servo", "PWM", None, 5.0), ("lora", "SPI", None, 3.3),
]

def detect_interface(name, description):
    text = (name + " " + (description or "")).lower()
    for hint, iface, addr, volt in INTERFACE_HINTS:
        if hint in text:
            return iface, addr, volt
    return None, None, None

def parse_suppliers(raw):
    if not raw:
        return []
    return [s.strip() for s in str(raw).replace(" / ", "/").split("/") if s.strip()]

PRIORITY_MAP = {"critical": "critical", "recommended": "recommended", "optional": "optional"}
STATUS_MAP = {
    "to order": "to_order", "in stock": "in_stock", "ordered": "ordered",
    "received": "received", "depleted": "depleted",
}

def norm_priority(raw, default="recommended"):
    return PRIORITY_MAP.get((raw or "").strip().lower(), default)

def norm_status(raw, default="to_order"):
    s = (raw or "").strip().lower()
    if s.startswith("unknown") or s.startswith("open decision"):
        return default
    return STATUS_MAP.get(s, default)

# ── row shape ────────────────────────────────────────────────────────────
# Every reader below produces a list of dicts with this shape.

def row(name, description="", category="", subcategory=None, morphology=None,
        material_spec=None, qty=1.0, unit="ea", unit_cost_cad=None,
        suppliers=None, interface_type=None, i2c_address=None, voltage_v=None,
        priority="recommended", status="to_order", notes=None, source_item_no=None):
    return dict(
        name=name.strip(), description=description, category=category,
        subcategory=subcategory, morphology=morphology, material_spec=material_spec,
        qty=qty, unit=unit, unit_cost_cad=unit_cost_cad,
        suppliers=suppliers or [], interface_type=interface_type,
        i2c_address=i2c_address, voltage_v=voltage_v,
        priority=priority, status=status, notes=notes, source_item_no=source_item_no,
    )

# ── readers ──────────────────────────────────────────────────────────────

def read_wrist_puck_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for i, r in enumerate(csv.DictReader(f)):
            if not r.get("name"):
                continue
            iface = r.get("interface_type") or None
            addr = r.get("i2c_address") or None
            volt = float(r["voltage_v"]) if r.get("voltage_v") else None
            if not iface:
                d_iface, d_addr, d_volt = detect_interface(r["name"], r.get("description", ""))
                iface, addr, volt = iface or d_iface, addr or d_addr, volt or d_volt
            rows.append(row(
                name=r["name"], description=r.get("description", ""),
                category=r.get("category", ""), subcategory=r.get("subcategory") or None,
                qty=float(r["qty"]) if r.get("qty") else 1.0,
                unit_cost_cad=float(r["unit_cost_cad"]) if r.get("unit_cost_cad") else None,
                suppliers=parse_suppliers(r.get("supplier_options")),
                interface_type=iface, i2c_address=addr, voltage_v=volt,
                priority=norm_priority(r.get("priority"), default="critical"),
                status=norm_status(r.get("status")),
                notes=r.get("notes") or None, source_item_no=i + 1,
            ))
    return rows

def read_ssf_xlsx(path):
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["BOM"]
    rows = []
    for r in ws.iter_rows(values_only=True):
        item_no = r[0]
        if not isinstance(item_no, (int, float)):
            continue
        name = str(r[1]).strip() if r[1] else ""
        if not name:
            continue
        description = str(r[2]).strip() if r[2] else ""
        category = str(r[3]).strip() if r[3] else ""
        morphology = str(r[4]).strip() if r[4] else None
        material = str(r[5]).strip() if r[5] else None
        qty = float(r[6]) if r[6] else 0
        unit = str(r[7]).strip() if r[7] else "ea"
        unit_cost = float(r[8]) if isinstance(r[8], (int, float)) else None
        supplier = str(r[10]).strip() if r[10] else ""
        iface, addr, volt = detect_interface(name, description)
        rows.append(row(
            name=name, description=description, category=category,
            morphology=morphology, material_spec=material, qty=qty, unit=unit,
            unit_cost_cad=unit_cost, suppliers=parse_suppliers(supplier),
            interface_type=iface, i2c_address=addr, voltage_v=volt,
            priority=norm_priority(str(r[11]) if r[11] else None),
            status=norm_status(str(r[12]) if r[12] else None),
            source_item_no=int(item_no),
        ))
    return rows

def read_morphogenesis_xlsx(path):
    wb = load_workbook(str(path), read_only=True, data_only=True)
    rows = []
    # "Hardware BOM": header at row idx 2 -> Component Name, Category, Qty,
    # Unit Price (CAD), Total (CAD), Supplier, Part # / Search Term, Notes
    ws = wb["Hardware BOM"]
    for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        name = str(r[0]).strip() if r[0] else ""
        if not name:
            continue
        category = str(r[1]).strip() if r[1] else ""
        if not category:
            # Section headers ("MOTION SYSTEM"), the "TOTAL HARDWARE BOM
            # COST" row, and the trailing pricing-disclaimer note all land
            # in column A with every other column blank — real line items
            # always carry a Category. Skip these rather than import them
            # as bogus zero-cost "parts".
            continue
        qty = float(r[2]) if isinstance(r[2], (int, float)) else 1.0
        unit_cost = float(r[3]) if isinstance(r[3], (int, float)) else None
        supplier = str(r[5]).strip() if r[5] else ""
        part_no = str(r[6]).strip() if r[6] else ""
        notes = str(r[7]).strip() if r[7] else None
        iface, addr, volt = detect_interface(name, notes or "")
        # morphology tags which sheet this row came from — without it, a part
        # name appearing in both sheets (e.g. "60mm Exhaust Fan 24V") would
        # collide under bom_line_items' (bom_id, part_id, morphology) upsert
        # key and silently drop one sheet's quantity.
        rows.append(row(
            name=name, description=part_no, category=category, qty=qty,
            morphology="Hardware BOM",
            unit_cost_cad=unit_cost, suppliers=parse_suppliers(supplier),
            interface_type=iface, i2c_address=addr, voltage_v=volt,
            notes=notes, source_item_no=i + 1,
        ))
    # "CF Upgrade Kit": header at row idx 2 -> Upgrade Item, Category,
    # Why It's Needed, Qty, Unit Price (CAD), Total (CAD), Supplier,
    # Part # / Search Term, Priority
    ws = wb["CF Upgrade Kit"]
    for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        name = str(r[0]).strip() if r[0] else ""
        if not name:
            continue
        category = str(r[1]).strip() if r[1] else ""
        if not category:
            # Same pattern as Hardware BOM above: "NOZZLES"/"HOTEND PATH"/
            # etc. section headers, "TOTAL CF UPGRADE KIT COST", and the
            # trailing priority-tier note row all have no Category.
            continue
        description = str(r[2]).strip() if r[2] else ""
        qty = float(r[3]) if isinstance(r[3], (int, float)) else 1.0
        unit_cost = float(r[4]) if isinstance(r[4], (int, float)) else None
        supplier = str(r[6]).strip() if r[6] else ""
        priority = str(r[8]).strip() if r[8] else None
        rows.append(row(
            name=name, description=description, category=category, qty=qty,
            morphology="CF Upgrade Kit",
            unit_cost_cad=unit_cost, suppliers=parse_suppliers(supplier),
            priority=norm_priority(priority), source_item_no=100 + i,
        ))
    return rows

LORA_SHEETS = ["Compute", "LoRa-Gateway", "Power", "Networking-Uplink", "Mechanical-Env"]

def read_lora_backbone_xlsx(path):
    wb = load_workbook(str(path), read_only=True, data_only=True)
    rows = []
    for sheet in LORA_SHEETS:
        ws = wb[sheet]
        # header at row idx 1 -> Line #, Part Name, SKU, Description, Qty,
        # Unit Price (USD), Extended (USD), Supplier, Supplier URL, Lead Time, Notes
        for r in ws.iter_rows(min_row=3, values_only=True):
            line_no = r[0]
            if not isinstance(line_no, (int, float)):
                continue
            name = str(r[1]).strip() if r[1] else ""
            if not name:
                continue
            sku = str(r[2]).strip() if r[2] else ""
            description = str(r[3]).strip() if r[3] else ""
            qty = float(r[4]) if isinstance(r[4], (int, float)) else 1.0
            unit_price_usd = float(r[5]) if isinstance(r[5], (int, float)) else None
            unit_cost_cad = round(unit_price_usd * FX_USD_TO_CAD, 4) if unit_price_usd else None
            supplier = str(r[7]).strip() if r[7] else ""
            supplier_url = str(r[8]).strip() if r[8] else ""
            lead_time = str(r[9]).strip() if r[9] else ""
            raw_notes = str(r[10]).strip() if r[10] else ""
            notes_parts = [p for p in [
                raw_notes,
                f"SKU: {sku}" if sku else None,
                f"lead time: {lead_time}" if lead_time else None,
                f"orig ${unit_price_usd:.2f} USD, {FX_SOURCE_NOTE}" if unit_price_usd else None,
                supplier_url or None,
            ] if p]
            iface, addr, volt = detect_interface(name, description)
            rows.append(row(
                name=name, description=description, category=sheet, qty=qty,
                morphology=sheet,  # same disambiguation reasoning as morphogenesis's two sheets
                unit_cost_cad=unit_cost_cad, suppliers=parse_suppliers(supplier),
                interface_type=iface, i2c_address=addr, voltage_v=volt,
                notes="; ".join(notes_parts) or None, source_item_no=int(line_no),
            ))
    return rows

# ── source registry ─────────────────────────────────────────────────────

SOURCES = {
    "wrist_puck": dict(
        bom_name="Wrist-Puck",
        reader=read_wrist_puck_csv,
        default_path="hardware/bom/wrist_puck_bom_v1_DRAFT.csv",
    ),
    "ssf": dict(
        bom_name="Sensor Species Family",
        reader=read_ssf_xlsx,
        default_path="hardware/bom/Sensor Species Family — Full Bill of Materials.xlsx",
    ),
    "morphogenesis": dict(
        bom_name="Morphogenesis",
        reader=read_morphogenesis_xlsx,
        default_path="hardware/bom/Morphogenesis BOM — Full Build, Filament, CF Upgrades & Supplies.xlsx",
    ),
    "lora_backbone": dict(
        bom_name="Backbone LoRa",
        reader=read_lora_backbone_xlsx,
        default_path="hardware/bom/Portenta–RPi5–Jetson LoRa Backbone BOM.xlsx",
    ),
}

# ── upserts ──────────────────────────────────────────────────────────────

def get_bom_source_id(cur, bom_name):
    cur.execute("SELECT id FROM bom_sources WHERE name = %s", (bom_name,))
    r = cur.fetchone()
    if not r:
        raise RuntimeError(
            f"bom_sources row '{bom_name}' not found — expected it seeded by schemas/003"
        )
    return str(r["id"])

def upsert_part(cur, r):
    cur.execute("""
        INSERT INTO parts (
            name, description, category, subcategory, material_spec,
            unit_type, unit_cost_cad, supplier_options, interface_type,
            i2c_address, voltage_v, notes
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (name) DO UPDATE SET
            description      = EXCLUDED.description,
            category         = EXCLUDED.category,
            subcategory      = COALESCE(EXCLUDED.subcategory, parts.subcategory),
            material_spec    = COALESCE(EXCLUDED.material_spec, parts.material_spec),
            unit_cost_cad    = COALESCE(EXCLUDED.unit_cost_cad, parts.unit_cost_cad),
            supplier_options = CASE WHEN EXCLUDED.supplier_options = '{}'
                                     THEN parts.supplier_options ELSE EXCLUDED.supplier_options END,
            interface_type   = COALESCE(EXCLUDED.interface_type, parts.interface_type),
            i2c_address      = COALESCE(EXCLUDED.i2c_address, parts.i2c_address),
            voltage_v        = COALESCE(EXCLUDED.voltage_v, parts.voltage_v),
            notes            = COALESCE(EXCLUDED.notes, parts.notes),
            updated_at       = NOW()
        RETURNING id
    """, (
        r["name"], r["description"], r["category"], r["subcategory"], r["material_spec"],
        r["unit"], r["unit_cost_cad"], r["suppliers"], r["interface_type"],
        r["i2c_address"], r["voltage_v"], r["notes"],
    ))
    return str(cur.fetchone()["id"])

def upsert_inventory(cur, part_id, r):
    qty_on_hand = r["qty"] if r["status"] == "in_stock" else 0
    qty_on_order = r["qty"] if r["status"] == "ordered" else 0
    cur.execute("""
        INSERT INTO inventory (part_id, qty_on_hand, qty_on_order, status)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (part_id) DO UPDATE SET
            status = CASE WHEN inventory.status = 'in_stock' THEN inventory.status
                           ELSE EXCLUDED.status END,
            last_updated = NOW()
    """, (part_id, qty_on_hand, qty_on_order, r["status"]))

def upsert_bom_line_item(cur, bom_id, part_id, r):
    # bom_line_items' UNIQUE (bom_id, part_id, morphology) does not dedup NULL
    # morphology against itself (Postgres treats each NULL as distinct) — only
    # SSF's rows carry a real morphology value; every other source leaves it
    # unset. Coalesce to '' so the ON CONFLICT target actually matches on
    # re-run instead of silently duplicating every non-SSF line item.
    morphology = r["morphology"] if r["morphology"] is not None else ""
    cur.execute("""
        INSERT INTO bom_line_items (bom_id, part_id, source_item_no, qty_required, morphology, priority, notes)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (bom_id, part_id, morphology) DO UPDATE SET
            qty_required = EXCLUDED.qty_required,
            priority     = EXCLUDED.priority
    """, (bom_id, part_id, r["source_item_no"], r["qty"], morphology, r["priority"], r["notes"]))

# ── main ─────────────────────────────────────────────────────────────────

def run_source(cur, key, path, dry_run):
    spec = SOURCES[key]
    rows = spec["reader"](path)
    print(f"\n{'='*70}\n{key} ({spec['bom_name']}) — {path}\n  {len(rows)} rows parsed")
    total = sum((r["unit_cost_cad"] or 0) * r["qty"] for r in rows)
    print(f"  estimated total: CAD ${total:,.2f}")
    if dry_run:
        for r in rows[:5]:
            print(f"    [{r['source_item_no']}] {r['name'][:45]:<45} "
                  f"${r['unit_cost_cad'] or 0:>8.2f}  {r['status']:10}  {r['priority']}")
        if len(rows) > 5:
            print(f"    ... and {len(rows) - 5} more")
        return len(rows), total

    bom_id = get_bom_source_id(cur, spec["bom_name"])
    for r in rows:
        part_id = upsert_part(cur, r)
        upsert_inventory(cur, part_id, r)
        upsert_bom_line_item(cur, bom_id, part_id, r)
    print(f"  upserted into bom_sources['{spec['bom_name']}']")
    return len(rows), total

def main():
    ap = argparse.ArgumentParser(description="Import Meridian hardware BOMs into the v2 parts schema")
    ap.add_argument("--source", choices=list(SOURCES) + ["all"], default="all")
    ap.add_argument("--bom-dir", default="hardware/bom", help="base dir for default file paths")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    keys = list(SOURCES) if args.source == "all" else [args.source]
    paths = {}
    for k in keys:
        default = SOURCES[k]["default_path"]
        # default_path is already relative to repo root; respect --bom-dir override
        # only for the filename portion.
        fname = Path(default).name
        paths[k] = Path(args.bom_dir) / fname
        if not paths[k].exists():
            print(f"ERROR: {paths[k]} not found", file=sys.stderr)
            sys.exit(1)

    if args.dry_run:
        print("DRY RUN — no database writes.\n")
        grand_total = 0
        for k in keys:
            n, total = run_source(None, k, paths[k], dry_run=True)
            grand_total += total
        print(f"\n{'='*70}\nGrand total across {len(keys)} source(s): CAD ${grand_total:,.2f}")
        return

    conn = db_connect()
    conn.autocommit = False
    cur = conn.cursor()
    try:
        grand_total = 0
        grand_rows = 0
        for k in keys:
            n, total = run_source(cur, k, paths[k], dry_run=False)
            grand_rows += n
            grand_total += total
        conn.commit()
        print(f"\n{'='*70}\n✓ committed. {grand_rows} line items across {len(keys)} source(s), "
              f"CAD ${grand_total:,.2f} estimated total.")

        cur.execute("SELECT COUNT(*) AS n FROM parts")
        print(f"  parts total in db:          {cur.fetchone()['n']}")
        cur.execute("SELECT COUNT(*) AS n FROM bom_line_items")
        print(f"  bom_line_items total in db: {cur.fetchone()['n']}")
    except Exception as e:
        conn.rollback()
        print(f"\nERROR (rolled back): {e}", file=sys.stderr)
        import traceback; traceback.print_exc()
        sys.exit(1)
    finally:
        cur.close()
        conn.close()

if __name__ == "__main__":
    main()
